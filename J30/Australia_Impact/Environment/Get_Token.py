import requests
import os
import re
import json
import time
import base64
from pathlib import Path

import pandas as pd
import urllib3
from openpyxl import load_workbook
from urllib3.exceptions import InsecureRequestWarning

from Australia_Impact.Inbound.Inbound_payload_generation.Inbound_State_Manager import StateManager


class Get_Token():

    # To get the bearer token
    # ----------------------------
    def __init__(self, env, plant, username=None, password=None):
        self.bearer_host_name = ''
        self.data = ''
        self.key = ''
        self.env = str(env or "").strip().lower()
        self.plant = str(plant or "").strip()
        self.state_manager = StateManager()

        if self.env == 'qa' and self.plant == '1081':
            self.bearer_host_name = "https://nikeaplawmqa1-auth.sce.manh.com/oauth/token"
        elif self.env == 'prod' and self.plant == '1081':
            self.bearer_host_name = "https://nikip3-auth.sce.manh.com/oauth/token"
        elif self.env == 'dev' and self.plant == '1081':
            self.bearer_host_name = "https://nikeaplawmdev1-auth.sce.manh.com/oauth/token"
        elif self.env == 'qa' and self.plant == '1093':
            self.bearer_host_name = "https://nikeaplawmqa1-auth.sce.manh.com/oauth/token"
        else:
            print("No proper Environment is passed to token class.")

        resolved_username = str(username or os.getenv("NIKE_USER_NAME", "")).strip()
        resolved_password = str(password or os.getenv("NIKE_PASSWORD", "")).strip()
        if not resolved_username or not resolved_password:
            resolved_username, resolved_password = self._load_credentials_from_inbound_master()

        if not resolved_username or not resolved_password:
            raise ValueError(
                "Missing credentials for token retrieval. Provide Username/Password in "
                "InboundMaster tab of Australia_Impact/Input_files/Inbound_worksheet.xlsx."
            )
        self.username = resolved_username

        self.payload = {
            "grant_type": "password",
            "username": resolved_username,
            "password": resolved_password
        }

        self.headers = {
            'Authorization': 'Basic b21uaWNvbXBvbmVudC4xLjAuMDpiNHM4cmdUeWc1NVhZTnVu',
            'Content-Type': 'application/x-www-form-urlencoded'
        }

    def _token_cache_key(self) -> str:
        return f"{self.env}|{self.plant}|{self.username}"

    @staticmethod
    def _decode_jwt_expiry(token: str) -> int:
        try:
            parts = token.split(".")
            if len(parts) < 2:
                return 0
            payload = parts[1]
            padding = "=" * (-len(payload) % 4)
            decoded = base64.urlsafe_b64decode(payload + padding).decode("utf-8")
            payload_obj = json.loads(decoded)
            return int(payload_obj.get("exp", 0) or 0)
        except Exception:
            return 0

    def _get_cached_token(self) -> str:
        cache_root = self.state_manager.get_value("token_cache", {})
        if not isinstance(cache_root, dict):
            return ""

        entry = cache_root.get(self._token_cache_key(), {})
        if not isinstance(entry, dict):
            return ""

        token = str(entry.get("access_token", "")).strip()
        expires_at = int(entry.get("expires_at", 0) or 0)
        now = int(time.time())
        # Keep a safety buffer before expiry.
        if token and expires_at > (now + 120):
            return token
        return ""

    def _save_cached_token(self, token: str):
        if not token:
            return
        cache_root = self.state_manager.get_value("token_cache", {})
        if not isinstance(cache_root, dict):
            cache_root = {}

        expires_at = self._decode_jwt_expiry(token)
        if not expires_at:
            expires_at = int(time.time()) + 1800

        cache_root[self._token_cache_key()] = {
            "access_token": token,
            "expires_at": int(expires_at),
            "updated_at": int(time.time()),
        }
        self.state_manager.set_value("token_cache", cache_root)

    @staticmethod
    def _is_masked_password(password: str) -> bool:
        password = str(password or "").strip()
        return bool(password) and re.fullmatch(r"\*+", password) is not None

    @staticmethod
    def _get_worksheet_path() -> Path:
        return Path(__file__).resolve().parent.parent / "Input_files" / "Inbound_worksheet.xlsx"

    def _get_stored_password_for_user(self, username: str) -> str:
        credentials = self.state_manager.get_value("inbound_master_credentials", {})
        if isinstance(credentials, dict):
            return str(credentials.get(username, "")).strip()
        return ""

    def _store_password_for_user(self, username: str, password: str):
        if not username or not password:
            return
        credentials = self.state_manager.get_value("inbound_master_credentials", {})
        if not isinstance(credentials, dict):
            credentials = {}
        credentials[username] = password
        self.state_manager.set_value("inbound_master_credentials", credentials)
        self.state_manager.set_value("inbound_master_last_user", username)

    def _mask_password_in_sheet(self, row_number: int, raw_password: str):
        if not raw_password or self._is_masked_password(raw_password):
            return
        worksheet_path = self._get_worksheet_path()
        if not worksheet_path.exists():
            return
        workbook = load_workbook(worksheet_path)
        if "InboundMaster" not in workbook.sheetnames:
            return
        sheet = workbook["InboundMaster"]
        password_col_idx = None
        for col_idx in range(1, sheet.max_column + 1):
            header_val = sheet.cell(row=1, column=col_idx).value
            if str(header_val or "").strip().lower() == "password":
                password_col_idx = col_idx
                break
        if password_col_idx is None:
            return
        sheet.cell(row=row_number, column=password_col_idx).value = "*" * max(len(raw_password), 8)
        workbook.save(worksheet_path)

    def _resolve_row_credentials(self, row: dict, row_number: int) -> tuple[str, str]:
        username = (
            str(row.get("Username", "")).strip()
            or str(row.get("username", "")).strip()
            or str(row.get("user_name", "")).strip()
            or str(row.get("User_Name", "")).strip()
        )
        raw_password = (
            str(row.get("Password", "")).strip()
            or str(row.get("password", "")).strip()
        )
        if not username:
            return "", ""
        if raw_password and not self._is_masked_password(raw_password):
            self._store_password_for_user(username, raw_password)
            self._mask_password_in_sheet(row_number=row_number, raw_password=raw_password)
            return username, raw_password
        if raw_password and self._is_masked_password(raw_password):
            return username, self._get_stored_password_for_user(username)
        return username, self._get_stored_password_for_user(username)

    def _load_credentials_from_inbound_master(self) -> tuple[str, str]:
        worksheet_path = self._get_worksheet_path()
        if not worksheet_path.exists():
            return "", ""
        try:
            df = pd.read_excel(worksheet_path, sheet_name="InboundMaster", dtype=str).fillna("")
        except Exception:
            return "", ""

        exact_rows = []
        fallback_rows = []
        for idx, row in df.iterrows():
            row_dict = row.to_dict()
            row_env = str(row_dict.get("Environment", "")).strip().lower()
            row_plant = str(row_dict.get("Plant", "")).strip()
            username, password = self._resolve_row_credentials(row_dict, idx + 2)
            if not username or not password:
                continue
            if row_env == self.env and row_plant == self.plant:
                exact_rows.append((username, password))
            fallback_rows.append((username, password))

        if exact_rows:
            return exact_rows[0]
        if fallback_rows:
            return fallback_rows[0]
        return "", ""

    def get_bearer(self):
        force_refresh = os.getenv("NIKE_FORCE_TOKEN_REFRESH", "").strip().lower() in {"1", "true", "yes", "y"}
        if not force_refresh:
            cached = self._get_cached_token()
            if cached:
                self.key = cached
                return self.key

        # SSL behavior can be controlled externally:
        # - NIKE_CA_BUNDLE=/path/to/cert.pem  (recommended)
        # - NIKE_DISABLE_SSL_VERIFY=true      (temporary workaround only)
        disable_ssl_verify = os.getenv("NIKE_DISABLE_SSL_VERIFY", "").strip().lower() in {"1", "true", "yes", "y"}
        ca_bundle = os.getenv("NIKE_CA_BUNDLE", "").strip() or os.getenv("REQUESTS_CA_BUNDLE", "").strip()
        if disable_ssl_verify:
            urllib3.disable_warnings(InsecureRequestWarning)
        verify = False if disable_ssl_verify else (ca_bundle if ca_bundle else True)

        response = requests.post(
            url=self.bearer_host_name,
            headers=self.headers,
            data=self.payload,
            verify=verify,
            timeout=30
        )
        response.raise_for_status()
        self.data = response.json()
        self.key = self.data["access_token"]
        self._save_cached_token(self.key)
        # token = "eyJhbGciOiJSUzI1NiIsInR5cCIgOiAiSldUIiwia2lkIiA6ICJhQ1I2NlhmWWZRb1hkbUltaHFOWkxVSTZFVWpxVHZQaWNLSnRtdG1DTzlzIn0.eyJleHAiOjE3ODI3NDUzNzcsImlhdCI6MTc4MjcwMjE3NywianRpIjoiMjhiNWVmMzgtMGNjZi00YjhlLWJjMGItZmIzMWE3YjJkMDZiIiwiaXNzIjoiaHR0cHM6Ly9uaWtlYXBsYXdtcHJvZDEtYXV0aC5zY2UubWFuaC5jb20vYXV0aC9yZWFsbXMvbWFhY3RpdmUiLCJzdWIiOiJ2Z2FuYTMiLCJ0eXAiOiJCZWFyZXIiLCJhenAiOiJvbW5pY29tcG9uZW50LjEuMC4wIiwic2Vzc2lvbl9zdGF0ZSI6IjNhMTIwNzFlLTAxZTItNDM4Mi04OTdhLTA0ZDhlZGJkMTQyYyIsImFjciI6IjEiLCJzY29wZSI6ImVtYWlsIGNvbXBvbmVudCBwcm9maWxlIG9wZW5pZCBvbW5pIiwic2lkIjoiM2ExMjA3MWUtMDFlMi00MzgyLTg5N2EtMDRkOGVkYmQxNDJjIiwidXNlck9yZ3MiOlsiQVBMQSIsIjEwODEiXSwiZW1haWxfdmVyaWZpZWQiOmZhbHNlLCJ1c2VyX25hbWUiOiJ2Z2FuYTMiLCJ1c2VyTG9jYXRpb25zIjpbeyJsb2NhdGlvbklkIjoiMTA4MSIsImxvY2F0aW9uVHlwZSI6ImR1bW15In1dLCJwcmVmZXJyZWRfdXNlcm5hbWUiOiJ2Z2FuYTMiLCJnaXZlbl9uYW1lIjoiVmlub3Roa3VtYXIiLCJsb2NhbGUiOiJlbiIsImF1dGhvcml0aWVzIjpbIlJPTEVfVVNFUiIsIlJPTEVfQVBMQS1QQUMgT3JnIExvZ2luIl0sInVzZXJUaW1lWm9uZSI6IkF1c3RyYWxpYS9NZWxib3VybmUiLCJlZGdlIjoiMCIsIm9yZ2FuaXphdGlvbiI6IkFQTEEiLCJhY2Nlc3N0b0FsbEJVcyI6ZmFsc2UsInRlbmFudElkIjoibmlraXNwcjMxbyIsIm5hbWUiOiJWaW5vdGhrdW1hciBHYW5hcGF0aHkiLCJ1c2VyRGVmYXVsdHMiOlt7ImRlZmF1bHRCdXNpbmVzc1VuaXQiOm51bGwsImRlZmF1bHRMb2NhdGlvbiI6IjEwODEiLCJkZWZhdWx0T3JnYW5pemF0aW9uIjoiMTA4MSJ9XSwiZmFtaWx5X25hbWUiOiJHYW5hcGF0aHkiLCJlbWFpbCI6IlZpbm90aGt1bWFyLkdhbmFwYXRoeUBuaWtlLmNvbSJ9.c4KGPY7jx7AQtJr3_4bPKbvywBWs5q-AJ9G-OW9CTfsnkqRwmJRAhD2FQQDlbi6jE_4ikMj6o2uS_gIjy1wZEeA1xUFFpIi0umM129CZKScZXMR1DJZU55fWxXfLlxdvLjZ92RBDsu3wi6EuNbkzFVbgp7Y21aX9ZiUIMKXMkdlgWa5G6ZcwnSgnOGgVn3xyupeeC_nSLXhik3NUX6ddxfXDZNPrso9ozHOVs1ea0E7Oq8DNpOP0pvP2iFkPEDGFxITWEfQn8-gEtmVCsHIa1EbYbTuM0DGa1V8Q_mR2W-9ihOUU-OgkhFYt5XJsRNqmpEoTgvZo3D7XTOAYl7pMtg"
        return self.key
        # return token

    # Bearer token request completed -------------------------------

if __name__ == '__main__':
    try:
        bearer = Get_Token(env='qa', plant='1093')
        get_token = bearer.get_bearer()
        print(get_token)
    except Exception as ex:
        print(f"Token fetch failed: {ex}")