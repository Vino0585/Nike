import requests
import pandas as pd
import logging
import os
from datetime import datetime
from openpyxl import load_workbook

from collections import defaultdict
from pathlib import Path
from Australia_Impact.Environment.Get_Token import Get_Token
from Australia_Impact.Environment.WM_Environment import AWM_Env
from Australia_Impact.Inbound.Inbound_payload_generation.ASN_Creation_Payload import Asn_Payload_Generator
from Australia_Impact.Inbound.Inbound_payload_generation.Execution_Report_Writer import (
    ExecutionReportWriter,
)

# Setup basic logging to provide better feedback than print()
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

SCRIPT_DIR = Path(__file__).resolve().parent
PROJECT_ROOT = SCRIPT_DIR.parent.parent
AUSTRALIA_IMPACT_ROOT = SCRIPT_DIR.parent

class ASN_Creation:

    @staticmethod
    def _get_ssl_verify_config():
        disable_ssl_verify = os.getenv("NIKE_DISABLE_SSL_VERIFY", "").strip().lower() in {"1", "true", "yes", "y"}
        ca_bundle = os.getenv("NIKE_CA_BUNDLE", "").strip() or os.getenv("REQUESTS_CA_BUNDLE", "").strip()
        return False if disable_ssl_verify else (ca_bundle if ca_bundle else True)

    @staticmethod
    def _normalize_asn_value(value: object) -> str:
        """Keep ASN values as text and preserve leading zeros."""
        raw = str(value or "").strip()
        if not raw:
            return ""
        # ASNs are expected to be 10-digit numeric strings in this AU flow.
        if raw.isdigit() and len(raw) < 10:
            return raw.zfill(10)
        return raw

    @staticmethod
    def _format_asnid_as_text(workbook_path: Path):
        """Force ASNID column to text format in MasterInput so Excel retains leading zeros."""
        try:
            workbook = load_workbook(workbook_path)
            if "MasterInput" not in workbook.sheetnames:
                workbook.close()
                return

            sheet = workbook["MasterInput"]
            headers = {}
            for col_idx in range(1, sheet.max_column + 1):
                header_val = str(sheet.cell(row=1, column=col_idx).value or "").strip()
                if header_val:
                    headers[header_val] = col_idx

            asn_col_idx = headers.get("ASNID")
            if not asn_col_idx:
                workbook.close()
                return

            for row_idx in range(2, sheet.max_row + 1):
                cell = sheet.cell(row=row_idx, column=asn_col_idx)
                normalized = ASN_Creation._normalize_asn_value(cell.value)
                if normalized:
                    cell.value = normalized
                cell.number_format = "@"

            workbook.save(workbook_path)
            workbook.close()
        except Exception as ex:
            logging.error(f"Failed to enforce text format for MasterInput.ASNID in {workbook_path}: {ex}")

    def create_asns(self):
        run_started_at = datetime.now()
        run_user = ""
        success_count = 0
        failure_count = 0
        step_records = []
        asn_gen = Asn_Payload_Generator()
        payload_packages = asn_gen.generate_payloads
        if not payload_packages:
            logging.error("No payloads were generated. Please check your Excel input and generator logic.")
            return

        payloads_by_env = defaultdict(list)
        for package in payload_packages:
            env = package.get('environment')
            payload = package.get('payload')
            if env and payload:
                payloads_by_env[env].append(payload)
            else:
                logging.error(f"WARNING: Skipping malformed package: {package}")

        # This list will collect data for the final report from ALL successful payloads
        extracted_report_data = []
        output_data = [] # This will hold one dictionary per successful payload
        verify = self._get_ssl_verify_config()

        for environment, payloads in payloads_by_env.items():
            logging.info(f"Processing {len(payloads)} Payloads for Environment: {environment.upper()}")
            if not payloads:
                logging.error(f"WARNING: Skipping empty payload list for environment {environment.upper()}.")
                continue

            try:
                plant_id_for_token = payloads[0].get('OrgId')
                if not plant_id_for_token:
                    logging.error("FATAL ERROR: Cannot get token. payload for {environment.upper()} is missing 'OrgId'")
                    continue

                token_handler = Get_Token(env=environment.lower(), plant=plant_id_for_token)
                bearer_token = token_handler.get_bearer()
                run_user = run_user or getattr(token_handler, "username", "")
                logging.info(f"Successfully retrieved token for {environment.upper()} environment.")

                env_handler = AWM_Env()

                for i, payload_to_send in enumerate(payloads):
                    try:
                        plant_id = payload_to_send['OrgId']
                        logging.info(
                            f"[{environment.upper()}] Processing Payload {i + 1}/{len(payloads)} for Plant {plant_id}")

                        env_handler.get_wm_host(host=environment.lower(), facility=str(plant_id))
                        url_value = env_handler.get_program_url(program="ASN_Creation")
                        if not url_value:
                            logging.error(
                                f"ERROR: Could not resolve endpoint URL for program ASN_Creation "
                                f"in {environment.upper()}/{plant_id}. Skipping payload {i + 1}."
                            )
                            continue
                        logging.info(f"Sending payload to URL: {url_value}")

                        headers = {
                            "content-type": "application/json",
                            "organization": str(plant_id),
                            "location": str(plant_id),
                            "authorization": 'Bearer ' + bearer_token
                        }

                        response = requests.post(
                            url=url_value,
                            headers=headers,
                            json=payload_to_send,
                            verify=verify,
                            timeout=30
                        )
                        response.raise_for_status()

                        response_data = response.json()
                        logging.info(f"Success: {response_data.get('success', 'N/A')}")
                        success_count += 1

                        # --- DATA COLLECTION FOR OUTPUT FILES ---
                        # This logic now runs only after a successful API call.
                        asn_id = payload_to_send.get('AsnId')
                        origin_facility = payload_to_send.get('OriginFacilityId')
                        lpn_list = payload_to_send.get('Lpn', [])
                        carrier_id = payload_to_send.get('CarrierId')

                        # 1. Data for the detailed report (ASN_Creation_Report.xlsx)
                        for lpn in lpn_list:
                            lpn_id = lpn.get('LpnId')
                            if lpn.get('LpnDetail'):
                                item_id = lpn['LpnDetail'][0].get('ItemId')
                                quantity = lpn['LpnDetail'][0].get('ShippedQuantity')

                                report_entry = {
                                    "PLANT": plant_id,
                                    "ENVN": environment,
                                    "ASN_ID": asn_id,
                                    "LPN_ID": lpn_id,
                                    "ITEM_ID": item_id,
                                    "QTY": quantity,
                                    "O_FACILITY": origin_facility,
                                    "CARRIER": carrier_id
                                }
                                extracted_report_data.append(report_entry)

                        # 2. Data for the input sheet (Inbound_worksheet.xlsx)
                        # This creates one row per successful payload.
                        current_lpns = [lpn.get('LpnId') for lpn in lpn_list if lpn.get('LpnId')]
                        formatted_lpn = ';'.join(current_lpns)
                        step_records.append(
                            {
                                "Environment": environment.upper(),
                                "Plant": plant_id,
                                "ASN_ID": asn_id,
                                "LPN_Count": len(current_lpns),
                                "LPNs": formatted_lpn,
                                "CarrierId": carrier_id,
                                "OriginFacilityId": origin_facility,
                                "ApiSuccess": response_data.get("success", "N/A"),
                            }
                        )

                        output_row = {
                            "PLANT": plant_id,
                            "ENVN": environment,
                            "ASN_ID": self._normalize_asn_value(asn_id),
                            "LPN_ID": formatted_lpn,
                            "Pre_Allocate": "Y",
                            "Failed": "N"
                        }
                        output_data.append(output_row)

                    except KeyError as e:
                        failure_count += 1
                        logging.error(f"ERROR: Could not process payload {i + 1}. Data is malformed. Missing key: {e}")
                    except requests.exceptions.RequestException as e:
                        failure_count += 1
                        logging.error(f"ERROR: API request failed for payload {i + 1}: {e}")
                        if e.response is not None:
                            logging.error(f"Status Code: {e.response.status_code}, Response: {e.response.text}")
                    except Exception as e:
                        failure_count += 1
                        logging.error(f"ERROR: An unexpected error occurred for payload {i + 1}: {e}")

            except Exception as e:
                failure_count += len(payloads)
                logging.error(f"FATAL ERROR: Could not process batch for environment {environment.upper()}. Error: {e}")

        # Generate the final report from ALL collected data
        if extracted_report_data:
            logging.info("Generating Report")
            try:
                report_df = pd.DataFrame(extracted_report_data)

                # Define the Output path.
                output_dir = AUSTRALIA_IMPACT_ROOT / "Output_files"
                output_dir.mkdir(parents=True, exist_ok=True)  # Just safe guaring.
                output_filepath = output_dir / "ASN_Creation_Report.xlsx"

                report_df.to_excel(output_filepath, index=False)
                logging.info(f"Successfully created report: {output_filepath}")
            except Exception as e:
                logging.error(f"Failed to create Excel report. Error: {e}")
        else:
            logging.info("No data was successfully processed to generate a report.")

        if output_data:
            logging.info("Generating input sheet from the create ASN output")
            try:
                report_df = pd.DataFrame(output_data)
                output_dir = AUSTRALIA_IMPACT_ROOT / "Input_files"
                output_dir.mkdir(parents=True, exist_ok=True)
                output_filepath = output_dir / "Inbound_worksheet.xlsx"
                asn_df = report_df.rename(columns={"PLANT": "Plant", "ENVN": "Environment", "ASN_ID": "ASNID",
                                                   "LPN_ID": "LPNID", "Pre_Allocate": 'Pre_Allocate',
                                                   "Failed": "Failed"})
                asn_df["ASNID"] = asn_df["ASNID"].apply(self._normalize_asn_value)

                if output_filepath.exists():
                    with pd.ExcelWriter(output_filepath, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                        asn_df.to_excel(writer, sheet_name='MasterInput', index=False)
                else:
                    with pd.ExcelWriter(output_filepath, engine='openpyxl', mode='w') as writer:
                        asn_df.to_excel(writer, sheet_name='MasterInput', index=False)

                self._format_asnid_as_text(output_filepath)

                logging.info(f"Successfully created multi-sheet report: {output_filepath}")

            except Exception as e:
                logging.error(f"ERROR: Failed to create multi-sheet Excel report. Error: {e}")

        else:
            logging.info("No data was successfully processed to generate an input sheet.")

        run_ended_at = datetime.now()
        report_path = ExecutionReportWriter().write_step_report(
            step_name="ASN Creation",
            run_user=run_user or os.getenv("USER", ""),
            started_at=run_started_at,
            ended_at=run_ended_at,
            status="SUCCESS" if success_count and not failure_count else ("PARTIAL" if success_count else "FAILED"),
            summary={
                "TotalPayloads": len(payload_packages),
                "SuccessfulPayloads": success_count,
                "FailedPayloads": failure_count,
            },
            records=step_records,
        )
        logging.info(f"Execution document generated: {report_path}")


if __name__ == '__main__':
    asn_create = ASN_Creation()
    asn_create.create_asns()